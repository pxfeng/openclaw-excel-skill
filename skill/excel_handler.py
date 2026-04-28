import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class ExcelHandler:
    def __init__(self, template_path, config):
        self.template_path = template_path
        self.config = config
        self.wb = None
        self.ws = None
        self.column_mapping = config.get('column_mapping', {})
        self.default_values = {}

    def load_template(self):
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"模板文件不存在: {self.template_path}")

        self.wb = load_workbook(self.template_path)
        self.ws = self.wb.active
        self.load_default_values()
        return self

    def load_default_values(self):
        header_row = self.find_header_row()
        default_row = header_row + 1

        for col_name, col_index in self.column_mapping.items():
            cell = self.ws.cell(row=default_row, column=col_index + 1)
            if cell.value is not None and str(cell.value).strip() != "":
                self.default_values[col_name] = cell.value

    def get_column_index(self, column_name):
        if column_name in self.column_mapping:
            return self.column_mapping[column_name]
        return None

    def find_header_row(self):
        for row in range(1, 10):
            for col in range(1, self.ws.max_column + 1):
                cell = self.ws.cell(row=row, column=col)
                if cell.value and "商品销售名称" in str(cell.value):
                    return row
        return 3

    def find_first_empty_row(self):
        for row in range(2, self.ws.max_row + 100):
            cell = self.ws.cell(row=row, column=1)
            if not cell.value:
                return row
        return self.ws.max_row + 1

    def write_data(self, data, use_defaults=True):
        if not self.ws:
            self.load_template()

        header_row = self.find_header_row()
        data_row = self.find_first_empty_row()

        for field_name, value in data.items():
            if value is None:
                continue

            col_index = self.get_column_index(field_name)
            if col_index is not None:
                cell = self.ws.cell(row=data_row, column=col_index + 1)
                cell.value = value

        if use_defaults:
            for col_name, default_val in self.get_all_defaults().items():
                col_index = self.get_column_index(col_name)
                if col_index is not None:
                    cell = self.ws.cell(row=data_row, column=col_index + 1)
                    if cell.value is None or cell.value == "":
                        cell.value = default_val

        return data_row

    def get_all_defaults(self):
        all_defaults = self.default_values.copy()
        
        fields = self.config.get('fields', {})
        for section, section_fields in fields.items():
            for field in section_fields:
                field_name = field.get('name')
                field_default = field.get('default')
                if field_name and field_default:
                    all_defaults[field_name] = field_default
        
        return all_defaults

    def get_default_values(self):
        return self.default_values.copy()

    def save(self, output_path):
        if not self.wb:
            raise Exception("请先加载模板")

        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)

        self.wb.save(output_path)
        return output_path

    def write_multiple_rows(self, data_list):
        rows = []
        for data in data_list:
            row = self.write_data(data)
            rows.append(row)
        return rows